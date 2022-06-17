import openpyxl

class ExcelGen:
    
    ex = openpyxl.Workbook()

    sheet = ex.active
    sheet.title = 'IMDB Top 250 Movies'

    sheet.append(['Search #1 and only one', ' ', ' ', ' '])
    sheet.append(['Movie Rank', 'Movie Name', 'Year of Realease', 'IMDB Rating'])

    def __init__(self, fn) -> None:
        self.fileName = fn


    def appendIt(self, sn:str, it:str):
        if sn not in self.ex.sheetnames:
            print("That sheet does not exist")
            return
        else:
            self.ex.active = self.ex[sn]
            self.sheet = self.ex.active
            self.sheet.append([it])

    def appendToSheet(self, sn:str, lst:list):
        if sn not in self.ex.sheetnames:
            print("That sheet does not exist")
            return
        else:
            self.ex.active = self.ex[sn]
            self.sheet = self.ex.active
            self.sheet.append(lst)

    def saveFile(self):
        try:
            self.ex.save(self.fileName)
        except Exception as e:
            print('\n',e)
            print("\nYou probably have the excel file opened, please close it and run this program again :)\n")
    
    def createSheet(self, sn:str):
        if sn not in self.ex.sheetnames:
            self.ex.create_sheet(sn)
        else:
            pass
